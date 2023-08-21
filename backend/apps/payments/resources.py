from datetime import date, datetime, timedelta
from io import BytesIO

from django.db.models import Sum
from django.http import HttpResponse
from django.contrib.auth import get_user_model
from django.utils.translation import gettext_lazy as _

from openpyxl import Workbook
from dateutil.relativedelta import relativedelta
from import_export import resources, fields
from import_export.widgets import ForeignKeyWidget

from utils.translations import get_verbose_name_by_name

from apps.clients.models import Client
from apps.projects.models import Project

from .models import ClientInvoice, ProjectExpense, ManagerBonus

User = get_user_model()


class ClientInvoiceResource(resources.ModelResource):
    invoice_number = fields.Field(
        column_name=get_verbose_name_by_name(ClientInvoice, 'invoice_number'),
        attribute='invoice_number',
    )
    legal_person = fields.Field(
        column_name=get_verbose_name_by_name(ClientInvoice, 'legal_person'),
        attribute='legal_person',
    )
    project = fields.Field(
        column_name=get_verbose_name_by_name(ClientInvoice, 'project'),
        attribute='project',
        widget=ForeignKeyWidget(Project, 'name'),
    )
    service = fields.Field(
        column_name=get_verbose_name_by_name(Project, 'service'),
        attribute='project__service',
        readonly=True,
    )
    payment_purpose = fields.Field(
        column_name=get_verbose_name_by_name(ClientInvoice, 'payment_purpose'),
        attribute='payment_purpose',
    )
    amount = fields.Field(
        column_name=get_verbose_name_by_name(ClientInvoice, 'amount'),
        attribute='amount',
    )
    billing_date = fields.Field(
        column_name=get_verbose_name_by_name(ClientInvoice, 'billing_date'),
        attribute='billing_date',
    )
    paid_date = fields.Field(
        column_name=get_verbose_name_by_name(ClientInvoice, 'paid_date'),
        attribute='paid_date',
    )
    close_date = fields.Field(
        column_name=get_verbose_name_by_name(ClientInvoice, 'close_date'),
        attribute='close_date',
    )
    comment = fields.Field(
        column_name=get_verbose_name_by_name(ClientInvoice, 'comment'),
        attribute='comment',
    )
    created_at = fields.Field(
        column_name=get_verbose_name_by_name(ClientInvoice, 'created_at'),
        attribute='created_at',
        readonly=True,
    )
    updated_at = fields.Field(
        column_name=get_verbose_name_by_name(ClientInvoice, 'updated_at'),
        attribute='updated_at',
        readonly=True,
    )
    deleted = fields.Field(
        column_name=get_verbose_name_by_name(ClientInvoice, 'deleted'),
        attribute='deleted',
        readonly=True,
    )
    deleted_comment = fields.Field(
        column_name=get_verbose_name_by_name(ClientInvoice, 'deleted_comment'),
        attribute='deleted_comment',
        readonly=True,
    )

    class Meta:
        model = ClientInvoice
        fields = (
            'id',
            'invoice_number',
            'billing_date',
            'project',
            'legal_person',
            'service',
            'payment_purpose',
            'amount',
            'paid_date',
            'close_date',
            'comment',
            'created_at',
            'updated_at',
            'deleted',
            'deleted_comment',
        )
        export_order = fields
        import_id_fields = ('invoice_number',)


class ProjectExpenseResource(resources.ModelResource):
    project = fields.Field(
        column_name=get_verbose_name_by_name(ProjectExpense, 'project'),
        attribute='project',
        widget=ForeignKeyWidget(Project, 'name'),
    )
    stock_market = fields.Field(
        column_name=get_verbose_name_by_name(ProjectExpense, 'stock_market'),
        attribute='stock_market',
    )
    month_request = fields.Field(
        column_name=get_verbose_name_by_name(ProjectExpense, 'month_request'),
        attribute='month_request',
    )
    budget_request = fields.Field(
        column_name=get_verbose_name_by_name(ProjectExpense, 'budget_request'),
        attribute='budget_request',
    )
    spent = fields.Field(
        column_name=get_verbose_name_by_name(ProjectExpense, 'spent'),
        attribute='spent',
    )
    commission = fields.Field(
        column_name=get_verbose_name_by_name(ProjectExpense, 'commission'),
        attribute='commission',
    )
    amount = fields.Field(
        column_name=get_verbose_name_by_name(ProjectExpense, 'amount'),
        attribute='amount',
    )
    project_comment = fields.Field(
        column_name=get_verbose_name_by_name(ProjectExpense, 'project_comment'),
        attribute='project_comment',
    )
    expense_comment = fields.Field(
        column_name=get_verbose_name_by_name(ProjectExpense, 'expense_comment'),
        attribute='expense_comment',
    )
    paid_date = fields.Field(
        column_name=get_verbose_name_by_name(ProjectExpense, 'paid_date'),
        attribute='paid_date',
    )
    deleted = fields.Field(
        column_name=get_verbose_name_by_name(ProjectExpense, 'deleted'),
        attribute='deleted',
        readonly=True,
    )
    deleted_comment = fields.Field(
        column_name=get_verbose_name_by_name(ProjectExpense, 'deleted_comment'),
        attribute='deleted_comment',
        readonly=True,
    )

    class Meta:
        model = ProjectExpense
        fields = (
            'id',
            'project',
            'stock_market',
            'month_request',
            'budget_request',
            'spent',
            'commission',
            'amount',
            'project_comment',
            'expense_comment',
            'paid_date',
            'deleted',
            'deleted_comment',
        )
        export_order = fields


class ManagerBonusResource(resources.ModelResource):
    manager = fields.Field(
        column_name=get_verbose_name_by_name(ManagerBonus, 'manager'),
        attribute='manager',
        widget=ForeignKeyWidget(User, 'username'),
    )
    client_invoice = fields.Field(
        column_name=get_verbose_name_by_name(ManagerBonus, 'client_invoice'),
        attribute='client_invoice',
        widget=ForeignKeyWidget(ClientInvoice, 'invoice_number'),
    )

    class Meta:
        model = ManagerBonus
        fields = (
            'id',
            'manager',
            'client_invoice',
        )

    def export(self, queryset=None, *args, **kwargs):
        """ Return response with xlsx file """

        if queryset is None:
            queryset = self.get_queryset()

        workbook = Workbook()
        workbook.remove(workbook.active)  # Удаляем пустую активную страницу

        start_date = kwargs.get('start_date')
        end_date = kwargs.get('end_date')

        if start_date and end_date:
            start_date = datetime.strptime(start_date, '%d.%m.%Y').date()
            end_date = datetime.strptime(end_date, '%d.%m.%Y').date()
        else:
            start_date = date(2000, 1, 1)
            end_date = date(2999, 1, 1)

        today = date.today()

        for manager_bonus in queryset:
            sheet = workbook.create_sheet(title=str(manager_bonus.manager))
            sheet.append([
                str(get_verbose_name_by_name(ClientInvoice, 'project')),
                str(get_verbose_name_by_name(ClientInvoice, 'legal_person')),
                str(get_verbose_name_by_name(ClientInvoice, 'invoice_number')),
                str(get_verbose_name_by_name(ClientInvoice, 'billing_date')),
                str(get_verbose_name_by_name(Client, 'legal_entity')),
                str(get_verbose_name_by_name(Project, 'service')),
                str(get_verbose_name_by_name(ClientInvoice, 'payment_purpose')),
                str(get_verbose_name_by_name(ClientInvoice, 'amount')),
                str(get_verbose_name_by_name(Project, 'expenses_amount')),
                str(get_verbose_name_by_name(Project, 'agent_commission_type')),
                str(get_verbose_name_by_name(Project, 'agent_commission')),
                str(_('Project month')),
                str(get_verbose_name_by_name(Project, 'manager_commission')),
                str(_('Bonus amount'))
            ])

            invoices = manager_bonus.client_invoice.filter(paid_date__range=(start_date, end_date))

            projects = []

            for invoice in invoices:
                # Заполняйте поля согласно вашей бизнес-логике
                project = invoice.project
                client = project.client
                service = project.service
                agent_commission_type = project.get_agent_commission_type_display()

                # Затем мы можем найти разницу между двумя датами
                difference = relativedelta(today, project.start_date)

                # И наконец, мы можем получить количество месяцев между этими двуми датами
                months_difference = (difference.years * 12 + difference.months) + 1

                manager_bonus_amount = 0

                invoice_amount = project.invoices.filter().aggregate(Sum('amount'))[
                    'amount__sum'] if project.invoices else 0

                if project not in projects:
                    expenses = project.expenses.filter().aggregate(Sum('amount'))[
                        'amount__sum'] if project.expenses else 0

                    if not expenses:
                        expenses = 0

                    projects.append(project)
                else:
                    expenses = 0

                if invoice_amount:
                    manager_bonus_amount = (invoice_amount - expenses) * (project.manager_commission / 100)

                sheet.append([
                    project.name,
                    invoice.legal_person,
                    invoice.invoice_number,
                    invoice.billing_date,
                    client.legal_entity,
                    service.name,
                    invoice.payment_purpose,
                    invoice_amount,
                    expenses,
                    agent_commission_type,
                    project.agent_commission_amount,
                    months_difference,
                    project.manager_commission,
                    manager_bonus_amount
                ])

        if not workbook.sheetnames:
            workbook.create_sheet(title='Empty')

        # Сохраняем в памяти, а не на диске
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Возвращаем файл пользователю через HTTP-ответ
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = 'attachment; filename=manager_bonus_export.xlsx'

        workbook.close()

        return response
